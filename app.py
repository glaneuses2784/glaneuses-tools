import streamlit as st
import pandas as pd
import requests
import statistics
from datetime import datetime

st.set_page_config(page_title="glaneuses 仕入れ判断ツール", page_icon="🛒", layout="wide")

# =====================================================
# 定数・設定
# =====================================================
APP_ID     = "KeikoFun-essearch-PRD-42e4e0b16-e0ae2b79"
EXCEL_FILE = "/Users/funahashikeiko/Desktop/glaneuses-company/経営・財務/仕入れ判断ツール_glaneuses.xlsx"

# 関税率 ── OC確定請求データ（2026年1〜4月 50件超）で実測・更新済み 2026-04-26
# 実効税率 = 実際のOC「米国関税」請求額 ÷ (売値USD × 155) で算出
# 食器は HTS サブコードにより 4〜29% と幅が広いため 20% を保守的中間値として使用
CUSTOMS_RATES = {
    "食器（磁器製）":              {"アメリカ": 0.200, "イギリス": 0.0, "EU（イタリア等）": 0.0},  # 実績4〜29%・平均17% / Royal Crown Derby等高級欧州品は28%超の場合あり
    "置物・フィギュリン（磁器製）":{"アメリカ": 0.140, "イギリス": 0.0, "EU（イタリア等）": 0.0},  # 実績11〜18%・平均14% ✅OC実測7件
    "花瓶・壺（磁器製）":          {"アメリカ": 0.190, "イギリス": 0.0, "EU（イタリア等）": 0.0},  # OC照合件数少・暫定維持
    "ガラス装飾・ボウル":          {"アメリカ": 0.130, "イギリス": 0.0, "EU（イタリア等）": 0.0},  # 実績5〜25%・平均13% ✅OC実測11件（Swarovski等で高め）
    "ガラスグラス（クリスタル）":  {"アメリカ": 0.140, "イギリス": 0.0, "EU（イタリア等）": 0.0},  # 実績13〜14%・平均14% ✅OC実測2件
    "ガラスグラス（非クリスタル）":{"アメリカ": 0.130, "イギリス": 0.0, "EU（イタリア等）": 0.0},  # 実績12〜13%・平均13% ✅OC実測4件
    "バッグ・革小物":              {"アメリカ": 0.160, "イギリス": 0.0, "EU（イタリア等）": 0.0},  # 実績11〜26%・平均16% ✅OC実測3件（HTSにより差大）
    "オルゴール・自動機械":        {"アメリカ": 0.100, "イギリス": 0.0, "EU（イタリア等）": 0.0},  # 実績5〜11%・平均8% / 安全サイドで10%維持
    "アンティーク（100年以上）":   {"アメリカ": 0.000, "イギリス": 0.0, "EU（イタリア等）": 0.0},  # 9706: Trump関税適用外
}
CATEGORIES   = list(CUSTOMS_RATES.keys())
DESTINATIONS = ["アメリカ", "イギリス", "EU（イタリア等）"]

# =====================================================
# データ読み込み（過去販売履歴）
# =====================================================
@st.cache_data
def load_sales_data():
    records = []
    try:
        file_2025 = "/Users/funahashikeiko/Desktop/glaneuses-tools/25年　利益管理表.xlsx"
        xl_2025 = pd.ExcelFile(file_2025)
        months = ['1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月']
        for month in months:
            if month in xl_2025.sheet_names:
                df = pd.read_excel(file_2025, sheet_name=month, header=0)
                n = len(df.columns)
                base = ['日付','商品名','販売金額USD','販売金額円','送料','関税','仕入れ値','粗利']
                df.columns = base[:n] + ['その他'] * max(0, n - len(base))
                for _, row in df.iterrows():
                    if pd.notna(row['商品名']) and pd.notna(row['販売金額USD']):
                        try:
                            usd = float(row['販売金額USD'])
                            if usd > 10000: continue
                        except: continue
                        jpy = row.get('販売金額円', None)
                        try: jpy = float(jpy) if pd.notna(jpy) else usd * 155
                        except: jpy = usd * 155
                        records.append({'年':'2025','商品名':str(row['商品名']).split('\n')[0][:80],
                            '販売金額USD':usd,'実質受取額円':jpy,'仕入れ値':row.get('仕入れ値',None),
                            '粗利':row.get('粗利',None)})
    except: pass
    try:
        file_2026 = "/Users/funahashikeiko/Desktop/glaneuses-tools/ebay メルカリ 仕入れ・売上管理 2026.xlsx"
        df_2026 = pd.read_excel(file_2026, sheet_name='eBay売上', header=0)
        for _, row in df_2026.iterrows():
            if pd.notna(row['商品名']) and pd.notna(row['販売金額(USD)']):
                try:
                    usd = float(row['販売金額(USD)'])
                    if usd > 10000: continue
                except: continue
                jpy = row.get('売上(円)', None)
                try: jpy = float(jpy) if pd.notna(jpy) else usd * 155
                except: jpy = usd * 155
                records.append({'年':'2026','商品名':str(row['商品名']).split('\n')[0][:80],
                    '販売金額USD':usd,'実質受取額円':jpy,'仕入れ値':row.get('仕入れ合計(円)',None),
                    '粗利':row.get('粗利(円)',None)})
    except: pass
    return pd.DataFrame(records)

# =====================================================
# eBay Finding API
# =====================================================
def search_ebay_sold(keywords: str, max_items: int = 100):
    url = "https://svcs.ebay.com/services/search/FindingService/v1"
    params = {
        "OPERATION-NAME": "findCompletedItems",
        "SERVICE-VERSION": "1.0.0",
        "SECURITY-APPNAME": APP_ID,
        "RESPONSE-DATA-FORMAT": "JSON",
        "keywords": keywords,
        "itemFilter(0).name": "SoldItemsOnly",
        "itemFilter(0).value": "true",
        "sortOrder": "EndTimeSoonest",
        "paginationInput.entriesPerPage": min(max_items, 100),
        "outputSelector(0)": "SellingStatus",
    }
    try:
        resp = requests.get(url, params=params, timeout=15)
        data = resp.json()
    except Exception as e:
        return None, f"接続エラー: {e}"

    if "errorMessage" in data:
        err = data["errorMessage"][0]["error"][0]
        err_id = err.get("errorId", ["?"])[0]
        if err_id == "10001":
            return None, "⚠️ 一時的なアクセス制限です。1〜2分後にもう一度お試しください。"
        return None, f"APIエラー (ID:{err_id})"

    try:
        items = data["findCompletedItemsResponse"][0]["searchResult"][0].get("item", [])
    except (KeyError, IndexError):
        return None, "データが見つかりませんでした"

    results = []
    for item in items:
        try:
            price    = float(item["sellingStatus"][0]["currentPrice"][0]["__value__"])
            currency = item["sellingStatus"][0]["currentPrice"][0]["@currencyId"]
            title    = item.get("title", [""])[0]
            results.append({"price": price, "currency": currency, "title": title})
        except: continue

    return results, None

def calc_profit(selling_price, destination, category, exchange_rate,
                buyer_shipping=49, oc_shipping=7000):
    total_fee = 0.15 + 0.02 + 0.02  # eBay + 広告 + Payoneer
    customs_rate = CUSTOMS_RATES.get(category, {}).get(destination, 0.0)
    revenue  = (selling_price + buyer_shipping) * exchange_rate
    fees     = revenue * total_fee
    customs  = selling_price * exchange_rate * customs_rate
    max_buy  = revenue - fees - customs - oc_shipping
    rec_buy  = max_buy / 1.3 if max_buy > 0 else 0
    return dict(revenue=revenue, fees=fees, customs=customs,
                max_buy=max_buy, rec_buy=rec_buy, customs_rate=customs_rate)

def save_research(keyword, category, destination, stats, selling_price, calc, note=""):
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        wb = load_workbook(EXCEL_FILE)
    except Exception as e:
        return False, str(e)

    sname = "リサーチ記録"
    if sname not in wb.sheetnames:
        ws = wb.create_sheet(sname)
        headers    = ["リサーチ日","キーワード","カテゴリ","送り先",
                      "90日売れ件数","最安値($)","最高値($)","平均値($)","中央値($)",
                      "想定販売価格($)","推奨仕入れ上限(円)","仕入れ結果","メモ"]
        col_widths = [12,30,20,16,12,10,10,10,10,16,18,12,28]
        thin = Side(style="thin", color="CCCCCC")
        for i,w in enumerate(col_widths,1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for col,h in enumerate(headers,1):
            c = ws.cell(1,col,h)
            c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
            c.fill      = PatternFill("solid", start_color="1F4E79")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = Border(left=thin,right=thin,top=thin,bottom=thin)
        ws.row_dimensions[1].height = 22
    else:
        ws = wb[sname]

    nr   = ws.max_row + 1
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin,right=thin,top=thin,bottom=thin)
    fg   = "FFFFFF" if nr % 2 == 0 else "EBF3FB"
    row  = [datetime.now().strftime("%Y-%m-%d"), keyword, category, destination,
            stats["count"], round(stats["min"]), round(stats["max"]),
            round(stats["avg"]), round(stats["median"]), round(selling_price),
            round(calc["rec_buy"]), "検討中", note]
    fmts = [None,None,None,None,"#,##0",'$#,##0','$#,##0','$#,##0','$#,##0',
            '$#,##0','#,##0;(#,##0);-',None,None]
    for col,(val,fmt) in enumerate(zip(row,fmts),1):
        c = ws.cell(nr,col,val)
        c.font      = Font(name="Arial", size=10,
                           bold=(col==11), color="0000FF" if col==11 else "000000")
        c.fill      = PatternFill("solid", start_color=fg)
        c.border    = bdr
        c.alignment = Alignment(horizontal="center" if col>4 else "left", vertical="center")
        if fmt: c.number_format = fmt

    wb.save(EXCEL_FILE)
    return True, ""

# =====================================================
# UI
# =====================================================
st.markdown("""
<style>
.main-header {
    background: linear-gradient(135deg, #1F4E79, #2E75B6);
    color: white; padding: 20px 30px; border-radius: 12px;
    margin-bottom: 20px;
}
.result-box {
    background: #f8f9fa; border-left: 4px solid #2E75B6;
    padding: 15px; border-radius: 0 8px 8px 0; margin: 10px 0;
}
.profit-ok   { background:#e8f5e9; border-left:4px solid #388e3c; padding:15px; border-radius:0 8px 8px 0; }
.profit-warn { background:#fff8e1; border-left:4px solid #f57c00; padding:15px; border-radius:0 8px 8px 0; }
.profit-ng   { background:#ffebee; border-left:4px solid #c62828; padding:15px; border-radius:0 8px 8px 0; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-header"><h2 style="margin:0">🛒 glaneuses 仕入れ判断ツール</h2>'
            '<p style="margin:4px 0 0 0; opacity:0.85">eBay相場リサーチ ／ 利益計算 ／ 過去販売履歴</p></div>',
            unsafe_allow_html=True)

tab1, tab2, tab3 = st.tabs(["🌐 eBay相場リサーチ", "💰 利益計算", "📦 過去販売履歴"])

# ─────────────────────────────────────────────────────
# TAB 1: eBay相場リサーチ
# ─────────────────────────────────────────────────────
with tab1:
    st.subheader("eBay 過去90日 売れ筋リサーチ")
    st.caption("キーワードを英語で入力するとeBayの直近の売れた価格を取得します")

    col_kw, col_btn = st.columns([4, 1])
    with col_kw:
        keyword = st.text_input("🔍 検索キーワード（英語）",
                                placeholder="例: Herend cup saucer  /  Royal Crown Derby plate",
                                label_visibility="collapsed")
    with col_btn:
        search_btn = st.button("検索", type="primary", use_container_width=True)

    if search_btn and keyword:
        with st.spinner(f"「{keyword}」を検索中..."):
            items, err = search_ebay_sold(keyword, 100)

        if err:
            st.error(err)
        elif not items:
            st.warning("データが見つかりませんでした。キーワードを変えてみてください。")
        else:
            prices = [i["price"] for i in items if i.get("currency") == "USD"]
            if not prices:
                st.warning("USD価格のデータがありませんでした。")
            else:
                stats = {"count": len(prices), "min": min(prices), "max": max(prices),
                         "avg": sum(prices)/len(prices), "median": statistics.median(prices)}

                st.success(f"✅ {stats['count']}件の売れた商品が見つかりました（過去90日）")

                c1,c2,c3,c4,c5 = st.columns(5)
                c1.metric("売れた件数",   f"{stats['count']}件")
                c2.metric("最安値",        f"${stats['min']:.0f}")
                c3.metric("最高値",        f"${stats['max']:.0f}")
                c4.metric("平均値",        f"${stats['avg']:.0f}")
                c5.metric("中央値 ★",     f"${stats['median']:.0f}")

                st.divider()
                st.caption("最近売れた商品（上位10件）")
                rows = [{"価格($)": f"${i['price']:.0f}", "商品タイトル": i['title'][:70]}
                        for i in items[:10]]
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

                st.divider()
                st.subheader("💰 仕入れ判断を計算する")

                col_a, col_b, col_c = st.columns(3)
                with col_a:
                    selling_price = st.number_input(
                        "想定販売価格（$）",
                        value=float(round(stats["median"])),
                        min_value=0.0, step=5.0,
                        help=f"中央値 ${stats['median']:.0f} が参考値です")
                with col_b:
                    category = st.selectbox("商品カテゴリ", CATEGORIES)
                with col_c:
                    destination = st.selectbox("送り先", DESTINATIONS)

                col_d, col_e = st.columns(2)
                with col_d:
                    exchange_rate = st.number_input("為替レート（円/$）", value=155, min_value=100, step=1)
                with col_e:
                    oc_shipping = st.number_input("OC送料実費（円）", value=7000, min_value=0, step=500)

                calc = calc_profit(selling_price, destination, category, exchange_rate,
                                   oc_shipping=oc_shipping)

                st.divider()
                c1,c2,c3,c4 = st.columns(4)
                c1.metric("売上（円）",       f"¥{calc['revenue']:,.0f}")
                c2.metric("手数料合計",        f"▲¥{calc['fees']:,.0f}", delta="eBay+広告+Payoneer")
                c3.metric(f"関税（{calc['customs_rate']*100:.1f}%）",
                          f"▲¥{calc['customs']:,.0f}")
                c4.metric("OC送料",            f"▲¥{oc_shipping:,}")

                profit = calc["max_buy"]
                rec    = calc["rec_buy"]

                if rec > 0:
                    css = "profit-ok"
                    icon = "✅"
                    msg = f"利益30%確保ライン：¥{rec:,.0f} 以下で仕入れれば利益30%以上"
                elif profit > 0:
                    css = "profit-warn"
                    icon = "⚠️"
                    msg = f"黒字ラインは ¥{profit:,.0f} 以下ですが利益は薄めです"
                else:
                    css = "profit-ng"
                    icon = "❌"
                    msg = "この価格帯では利益が出にくい状況です。値上げまたはEU向け販売を検討してください"

                st.markdown(f'<div class="{css}"><b>{icon} 推奨仕入れ上限：¥{rec:,.0f}</b><br>{msg}</div>',
                            unsafe_allow_html=True)

                st.divider()
                with st.expander("📥 Excelのリサーチ記録に保存する"):
                    note = st.text_input("メモ（省略可）", placeholder="例: オークションで見かけた / 状態良好")
                    if st.button("保存する", type="primary"):
                        ok, errmsg = save_research(keyword, category, destination,
                                                   stats, selling_price, calc, note)
                        if ok:
                            st.success("✅ 仕入れ判断ツールの「リサーチ記録」シートに保存しました")
                        else:
                            st.error(f"保存に失敗しました: {errmsg}")

# ─────────────────────────────────────────────────────
# TAB 2: 利益計算（関税対応版）
# ─────────────────────────────────────────────────────
with tab2:
    st.subheader("仕入れ前の利益シミュレーション")
    st.caption("関税・広告費・Payoneer手数料をすべて含んだ正確な計算です")

    col1, col2 = st.columns(2)
    with col1:
        item_name2    = st.text_input("商品名", placeholder="例: Royal Crown Derby プレート")
        buy_price2    = st.number_input("仕入れ価格（円）", min_value=0, step=500)
        ebay_price2   = st.number_input("eBay販売価格（$）", min_value=0.0, step=5.0)
        ship_income2  = st.number_input("受取送料（$）", min_value=0.0, value=49.0, step=1.0)

    with col2:
        category2    = st.selectbox("商品カテゴリ", CATEGORIES, key="cat2")
        destination2 = st.selectbox("送り先", DESTINATIONS, key="dest2")
        exchange2    = st.number_input("為替レート（円/$）", min_value=100, value=155, step=1, key="ex2")
        oc_ship2     = st.number_input("OC送料実費（円）", min_value=0, value=7000, step=500, key="oc2")

    if st.button("計算する", type="primary", key="calc2"):
        if ebay_price2 > 0 and buy_price2 > 0:
            calc2 = calc_profit(ebay_price2, destination2, category2, exchange2,
                                buyer_shipping=ship_income2, oc_shipping=oc_ship2)

            profit2 = calc2["revenue"] - calc2["fees"] - calc2["customs"] - oc_ship2 - buy_price2

            st.divider()
            c1,c2,c3,c4 = st.columns(4)
            c1.metric("売上（円）",   f"¥{calc2['revenue']:,.0f}")
            c2.metric("手数料合計",   f"▲¥{calc2['fees']:,.0f}")
            c3.metric(f"関税（{calc2['customs_rate']*100:.1f}%）",
                      f"▲¥{calc2['customs']:,.0f}")
            c4.metric("利益",         f"¥{profit2:,.0f}",
                      delta="利益あり ✅" if profit2 > 0 else "赤字 ❌")

            margin = profit2 / buy_price2 * 100 if buy_price2 > 0 else 0
            if profit2 >= buy_price2 * 0.3:
                st.success(f"✅ {item_name2}　利益 ¥{profit2:,.0f}（利益率 {margin:.0f}%）")
            elif profit2 > 0:
                st.warning(f"⚠️ {item_name2}　利益 ¥{profit2:,.0f}（利益率 {margin:.0f}%）やや薄め")
            else:
                st.error(f"❌ {item_name2}　損失 ¥{abs(profit2):,.0f}　見送り推奨")
        else:
            st.warning("仕入れ価格とeBay販売価格を入力してください。")

# ─────────────────────────────────────────────────────
# TAB 3: 過去販売履歴
# ─────────────────────────────────────────────────────
with tab3:
    st.subheader("過去の自分の販売履歴を検索")
    st.caption("2025〜2026年の販売データから類似商品の実績を確認できます")

    df = load_sales_data()

    col_s, col_n = st.columns([3, 2])
    with col_s:
        kw3 = st.text_input("キーワード（スペース区切りで絞り込み）",
                            placeholder="例: Richard Ginori soup bowl",
                            label_visibility="collapsed")
    with col_n:
        st.info("💡 2025年は送料無料、2026年は買い手負担のため実質受取額(円)で比較するのが正確です")

    if kw3:
        keywords3 = kw3.strip().split()
        result3   = df.copy()
        for kw in keywords3:
            result3 = result3[result3['商品名'].str.lower().str.contains(kw.lower(), na=False)]

        if result3.empty:
            st.warning(f"「{kw3}」の販売履歴は見つかりませんでした。")
        else:
            p_jpy  = pd.to_numeric(result3['実質受取額円'], errors='coerce').dropna()
            p_usd  = pd.to_numeric(result3['販売金額USD'], errors='coerce').dropna()
            profs  = pd.to_numeric(result3['粗利'], errors='coerce').dropna()
            profs  = profs[profs != 0]

            c1,c2,c3,c4 = st.columns(4)
            c1.metric("件数",            f"{len(result3)}件")
            c2.metric("平均受取額(円)",  f"¥{p_jpy.mean():,.0f}")
            c3.metric("平均販売価格($)", f"${p_usd.mean():.0f}")
            if not profs.empty:
                c4.metric("平均粗利",    f"¥{profs.mean():,.0f}")

            st.divider()
            disp = result3[['年','商品名','販売金額USD','実質受取額円','仕入れ値','粗利']].copy()
            disp['販売金額USD']  = pd.to_numeric(disp['販売金額USD'],  errors='coerce').apply(lambda x: f"${x:.0f}"    if pd.notna(x) else "ー")
            disp['実質受取額円'] = pd.to_numeric(disp['実質受取額円'], errors='coerce').apply(lambda x: f"¥{x:,.0f}"  if pd.notna(x) else "ー")
            disp['仕入れ値']     = pd.to_numeric(disp['仕入れ値'],     errors='coerce').apply(lambda x: f"¥{x:,.0f}"  if pd.notna(x) and x>0 else "ー")
            disp['粗利']         = pd.to_numeric(disp['粗利'],         errors='coerce').apply(lambda x: f"¥{x:,.0f}"  if pd.notna(x) and x!=0 else "ー")
            disp.columns = ['年','商品名','販売価格($)','実質受取額(円)','仕入れ値','粗利']
            st.dataframe(disp.reset_index(drop=True), use_container_width=True, height=400)

